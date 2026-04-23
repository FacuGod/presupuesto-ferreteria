from __future__ import annotations

import cgi
import json
import mimetypes
import socket
import sys
import tempfile
import threading
import webbrowser
from datetime import datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from presupuestos_ferreteria import (
    PRICE_MODE_TO_FIELD,
    auto_quote_number,
    load_products_from_excel,
    normalize_text,
    parse_positive_number,
)


APP_TITLE = "Presupuestos Ferretería"


class AppState:
    def __init__(self) -> None:
        self.products: list[dict] = []
        self.current_price_file = ""
        self.lock = threading.Lock()

    def set_products(self, products: list[dict], file_name: str) -> None:
        with self.lock:
            self.products = products
            self.current_price_file = file_name

    def snapshot(self) -> dict:
        with self.lock:
            return {
                "products": [serializable_product(product) for product in self.products],
                "current_price_file": self.current_price_file,
                "quote_number": auto_quote_number(),
                "date": datetime.now().strftime("%d/%m/%Y"),
            }


STATE = AppState()


def serializable_product(product: dict) -> dict:
    return {
        "codigo_interno": product.get("codigo_interno", ""),
        "producto": product.get("producto", ""),
        "descripcion_adicional": product.get("descripcion_adicional", ""),
        "marca": product.get("marca", ""),
        "indice": product.get("indice", ""),
        "precio_lista_unitario": product.get("precio_lista_unitario", 0.0),
        "precio_lista_neto_unitario": product.get("precio_lista_neto_unitario", 0.0),
        "precio_lista_neto_100_unid": product.get("precio_lista_neto_100_unid", 0.0),
        "search_text": product.get("search_text", ""),
    }


def safe_file_name(value: str) -> str:
    cleaned = "".join(ch for ch in value if ch.isalnum() or ch in ("-", "_", ".")).strip()
    return cleaned or auto_quote_number()


def export_quote_excel(file_path: str | Path, payload: dict) -> None:
    quote_items = payload.get("items", [])
    if not quote_items:
        raise ValueError("Agregá al menos un producto antes de guardar el presupuesto.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=15, color="0F172A")
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    total_fill = PatternFill("solid", fgColor="BFDBFE")
    thin_side = Side(style="thin", color="CBD5E1")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws["A1"] = "PRESUPUESTO"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")

    metadata = [
        ("Número", payload.get("quote_number", "")),
        ("Fecha", payload.get("date", "")),
        ("Cliente", payload.get("client_name", "")),
        ("Teléfono", payload.get("client_phone", "")),
        ("Dirección / Observación", payload.get("client_address", "")),
        ("Lista de precios usada", payload.get("current_price_file", "")),
    ]
    row_cursor = 3
    for label, value in metadata:
        ws.cell(row=row_cursor, column=1, value=label).font = bold
        ws.cell(row=row_cursor, column=2, value=value)
        row_cursor += 1

    row_cursor += 1
    headers = ["#", "Código", "Detalle", "Cantidad", "Precio unitario", "Subtotal"]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_cursor, column=col_index, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    data_start = row_cursor + 1
    for item_index, item in enumerate(quote_items, start=1):
        quantity = float(item.get("cantidad", 0) or 0)
        unit_price = float(item.get("precio_unitario", 0) or 0)
        subtotal = quantity * unit_price
        row_idx = data_start + item_index - 1
        values = [
            item_index,
            item.get("codigo", ""),
            item.get("detalle", ""),
            quantity,
            unit_price,
            subtotal,
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_index, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col_index in {4, 5, 6} else "left")

    for row_idx in range(data_start, data_start + len(quote_items)):
        ws.cell(row=row_idx, column=4).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=5).number_format = "$ #,##0.00"
        ws.cell(row=row_idx, column=6).number_format = "$ #,##0.00"

    subtotal = sum(float(item.get("cantidad", 0) or 0) * float(item.get("precio_unitario", 0) or 0) for item in quote_items)
    discount_pct = parse_positive_number(str(payload.get("discount", 0)))
    iva_pct = parse_positive_number(str(payload.get("iva", 0)))
    extra_cost = parse_positive_number(str(payload.get("extra", 0)))
    discount_amount = subtotal * discount_pct / 100
    base = subtotal - discount_amount
    iva_amount = base * iva_pct / 100
    total = base + iva_amount + extra_cost

    total_start = data_start + len(quote_items) + 2
    totals = [
        ("Subtotal", subtotal),
        (f"Descuento ({discount_pct:.2f}%)", discount_amount),
        (f"IVA ({iva_pct:.2f}%)", iva_amount),
        ("Recargo / Flete", extra_cost),
        ("TOTAL", total),
    ]
    for offset, (label, value) in enumerate(totals):
        row_idx = total_start + offset
        label_cell = ws.cell(row=row_idx, column=5, value=label)
        amount_cell = ws.cell(row=row_idx, column=6, value=value)
        label_cell.font = bold
        amount_cell.font = bold if label == "TOTAL" else Font(bold=False)
        amount_cell.number_format = "$ #,##0.00"
        label_cell.border = border
        amount_cell.border = border
        if label == "TOTAL":
            label_cell.fill = total_fill
            amount_cell.fill = total_fill

    widths = {"A": 8, "B": 16, "C": 60, "D": 12, "E": 18, "F": 18}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows():
        for cell in row:
            if cell.column == 3:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(file_path)


HTML = r"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Presupuestos Ferretería</title>
  <link rel="stylesheet" href="/app.css">
</head>
<body>
  <div class="app-shell">
    <header class="topbar">
      <div>
        <p class="eyebrow">Ferretería</p>
        <h1>Presupuestos</h1>
        <p class="subtitle">Busca productos, arma el pedido y exporta el Excel final.</p>
      </div>
      <div class="top-actions">
        <input id="priceFile" type="file" accept=".xlsx,.xlsm,.xltx,.xltm" hidden>
        <button class="btn accent" id="pickFile">Cargar lista</button>
        <button class="btn primary" id="saveQuote">Exportar Excel</button>
        <button class="btn ghost" id="clearQuote">Limpiar</button>
        <button class="icon-btn" id="closeApp" title="Cerrar aplicación">×</button>
      </div>
    </header>

    <section class="status-strip">
      <span id="statusText">Cargá una lista de precios para empezar.</span>
      <strong id="productCount">0 productos</strong>
    </section>

    <main class="layout">
      <section class="workspace">
        <div class="quote-info panel">
          <label>Número<input id="quoteNumber" type="text"></label>
          <label>Fecha<input id="quoteDate" type="text"></label>
          <label>Cliente<input id="clientName" type="text" placeholder="Nombre del cliente"></label>
          <label>Teléfono<input id="clientPhone" type="text" placeholder="Contacto"></label>
          <label class="wide">Dirección / Observación<input id="clientAddress" type="text" placeholder="Notas para este presupuesto"></label>
        </div>

        <section class="panel product-panel">
          <div class="panel-head">
            <div>
              <p class="eyebrow">Catálogo</p>
              <h2>Productos</h2>
            </div>
            <div class="controls">
              <input id="search" class="search" type="search" placeholder="Buscar por código, producto, marca o índice">
              <input id="quantity" class="quantity" type="number" min="0.01" step="0.01" value="1">
              <select id="priceMode"></select>
            </div>
          </div>
          <div class="table-wrap">
            <table>
              <thead>
                <tr>
                  <th>Código</th>
                  <th>Producto</th>
                  <th>Marca</th>
                  <th>Lista</th>
                  <th>Neto</th>
                  <th>100 unid.</th>
                  <th></th>
                </tr>
              </thead>
              <tbody id="productsBody"></tbody>
            </table>
          </div>
        </section>
      </section>

      <aside class="side">
        <section class="panel cart-panel">
          <div class="panel-head">
            <div>
              <p class="eyebrow">Pedido</p>
              <h2>Presupuesto actual</h2>
            </div>
          </div>
          <div id="emptyQuote" class="empty">Todavía no agregaste productos.</div>
          <div class="quote-list" id="quoteList"></div>
        </section>

        <section class="panel totals-panel">
          <div class="calc-grid">
            <label>Descuento %<input id="discount" type="number" min="0" step="0.01" value="0"></label>
            <label>IVA %<input id="iva" type="number" min="0" step="0.01" value="21"></label>
            <label>Recargo / Flete<input id="extra" type="number" min="0" step="0.01" value="0"></label>
          </div>
          <dl class="totals">
            <div><dt>Subtotal</dt><dd id="subtotal">$ 0,00</dd></div>
            <div><dt>Descuento</dt><dd id="discountAmount">$ 0,00</dd></div>
            <div><dt>IVA</dt><dd id="ivaAmount">$ 0,00</dd></div>
          </dl>
          <div class="grand-total">
            <span>Total</span>
            <strong id="total">$ 0,00</strong>
          </div>
        </section>
      </aside>
    </main>
  </div>
  <script src="/app.js"></script>
</body>
</html>
"""


CSS = r""":root {
  color-scheme: light;
  --ink: #102033;
  --muted: #66758a;
  --line: #d9e3ef;
  --surface: rgba(255, 255, 255, 0.9);
  --blue: #2457f5;
  --blue-dark: #1537a6;
  --orange: #ff7a1a;
  --teal: #00a88f;
  --rose: #e13d5c;
  --soft-blue: #eaf1ff;
  --shadow: 0 18px 50px rgba(28, 49, 86, 0.14);
}

* { box-sizing: border-box; }

body {
  margin: 0;
  min-width: 1100px;
  background:
    radial-gradient(circle at top left, rgba(36, 87, 245, 0.18), transparent 28rem),
    radial-gradient(circle at 85% 10%, rgba(255, 122, 26, 0.16), transparent 24rem),
    linear-gradient(135deg, #f8fbff 0%, #eef5f3 100%);
  color: var(--ink);
  font-family: Inter, "Segoe UI", system-ui, -apple-system, sans-serif;
}

button, input, select { font: inherit; }

.app-shell { padding: 22px; }

.topbar {
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 24px;
  padding: 24px 26px;
  border-radius: 28px;
  background: linear-gradient(135deg, #102033 0%, #173fb8 58%, #00a88f 100%);
  color: #fff;
  box-shadow: var(--shadow);
}

.eyebrow {
  margin: 0 0 6px;
  color: #ffbf75;
  font-size: 0.76rem;
  font-weight: 800;
  letter-spacing: 0.12em;
  text-transform: uppercase;
}

h1, h2 { margin: 0; letter-spacing: 0; }
h1 { font-size: 2.2rem; }
h2 { font-size: 1.15rem; }

.subtitle { margin: 8px 0 0; color: rgba(255, 255, 255, 0.78); }

.top-actions {
  display: flex;
  align-items: center;
  gap: 10px;
  flex-wrap: wrap;
  justify-content: flex-end;
}

.btn, .icon-btn {
  border: 0;
  cursor: pointer;
  transition: transform 160ms ease, box-shadow 160ms ease, background 160ms ease;
}

.btn {
  min-height: 42px;
  padding: 0 16px;
  border-radius: 999px;
  font-weight: 800;
  color: #fff;
}

.btn:hover, .icon-btn:hover { transform: translateY(-1px); }
.btn.primary { background: var(--blue); box-shadow: 0 12px 25px rgba(36, 87, 245, 0.28); }
.btn.accent { background: var(--orange); box-shadow: 0 12px 25px rgba(255, 122, 26, 0.28); }
.btn.ghost { background: rgba(255, 255, 255, 0.16); color: #fff; }

.icon-btn {
  width: 42px;
  height: 42px;
  border-radius: 50%;
  background: rgba(255, 255, 255, 0.16);
  color: #fff;
  font-size: 1.5rem;
  line-height: 1;
}

.status-strip {
  display: flex;
  justify-content: space-between;
  gap: 16px;
  margin: 16px 0;
  padding: 13px 18px;
  border: 1px solid rgba(36, 87, 245, 0.14);
  border-radius: 18px;
  background: rgba(255, 255, 255, 0.72);
  color: var(--muted);
  backdrop-filter: blur(14px);
}

.status-strip strong { color: var(--blue-dark); }

.layout {
  display: grid;
  grid-template-columns: minmax(680px, 1fr) 430px;
  gap: 18px;
  align-items: start;
}

.workspace, .side {
  display: flex;
  flex-direction: column;
  gap: 18px;
}

.panel {
  border: 1px solid rgba(217, 227, 239, 0.9);
  border-radius: 24px;
  background: var(--surface);
  box-shadow: var(--shadow);
  backdrop-filter: blur(18px);
}

.quote-info {
  display: grid;
  grid-template-columns: 1fr 0.75fr 1.3fr 0.95fr;
  gap: 14px;
  padding: 18px;
}

label {
  display: grid;
  gap: 7px;
  color: var(--muted);
  font-size: 0.82rem;
  font-weight: 800;
}

label.wide {
  grid-column: 1 / -1;
}

input, select {
  width: 100%;
  min-height: 42px;
  border: 1px solid var(--line);
  border-radius: 14px;
  background: #fff;
  color: var(--ink);
  outline: none;
  padding: 0 13px;
}

input:focus, select:focus {
  border-color: var(--blue);
  box-shadow: 0 0 0 4px rgba(36, 87, 245, 0.12);
}

.product-panel, .cart-panel, .totals-panel { padding: 18px; }

.panel-head {
  display: flex;
  align-items: end;
  justify-content: space-between;
  gap: 16px;
  margin-bottom: 14px;
}

.controls {
  display: grid;
  grid-template-columns: minmax(280px, 1fr) 92px 220px;
  gap: 10px;
  flex: 1;
  max-width: 760px;
}

.table-wrap {
  overflow: auto;
  max-height: calc(100vh - 380px);
  min-height: 330px;
  border: 1px solid var(--line);
  border-radius: 18px;
  background: #fff;
}

table {
  width: 100%;
  border-collapse: collapse;
  font-size: 0.92rem;
}

thead {
  position: sticky;
  top: 0;
  z-index: 1;
  background: #102033;
  color: #fff;
}

th, td {
  padding: 12px 13px;
  border-bottom: 1px solid #edf2f7;
  text-align: left;
  vertical-align: top;
}

th { font-size: 0.78rem; text-transform: uppercase; letter-spacing: 0.06em; }
td.money, th.money { text-align: right; white-space: nowrap; }
tbody tr:nth-child(even) { background: #f8fbff; }
tbody tr:hover { background: #eaf1ff; }

.add-row {
  min-width: 34px;
  height: 34px;
  border: 0;
  border-radius: 50%;
  background: var(--teal);
  color: #fff;
  cursor: pointer;
  font-weight: 900;
}

.quote-list {
  display: grid;
  gap: 10px;
  max-height: calc(100vh - 470px);
  min-height: 230px;
  overflow: auto;
  padding-right: 4px;
}

.quote-item {
  display: grid;
  grid-template-columns: 1fr auto;
  gap: 12px;
  padding: 13px;
  border: 1px solid var(--line);
  border-radius: 18px;
  background: #fff;
}

.quote-item strong {
  display: block;
  margin-bottom: 5px;
}

.quote-meta {
  color: var(--muted);
  font-size: 0.86rem;
}

.remove {
  width: 34px;
  height: 34px;
  border: 0;
  border-radius: 50%;
  background: #ffe4ea;
  color: var(--rose);
  cursor: pointer;
  font-weight: 900;
}

.empty {
  display: none;
  align-items: center;
  justify-content: center;
  min-height: 180px;
  border: 1px dashed var(--line);
  border-radius: 18px;
  color: var(--muted);
  background: #fff;
}

.empty.visible { display: flex; }

.calc-grid {
  display: grid;
  grid-template-columns: 1fr 1fr 1fr;
  gap: 10px;
}

.totals {
  display: grid;
  gap: 8px;
  margin: 16px 0;
}

.totals div {
  display: flex;
  justify-content: space-between;
  gap: 14px;
  padding: 10px 0;
  border-bottom: 1px solid #edf2f7;
}

dt { color: var(--muted); font-weight: 800; }
dd { margin: 0; font-weight: 900; }

.grand-total {
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 14px;
  padding: 18px;
  border-radius: 20px;
  background: linear-gradient(135deg, var(--blue) 0%, var(--teal) 100%);
  color: #fff;
}

.grand-total span {
  color: rgba(255, 255, 255, 0.82);
  font-weight: 800;
  text-transform: uppercase;
  letter-spacing: 0.08em;
}

.grand-total strong {
  font-size: 1.75rem;
}

.toast {
  position: fixed;
  right: 24px;
  bottom: 24px;
  max-width: 420px;
  padding: 14px 16px;
  border-radius: 16px;
  background: #102033;
  color: #fff;
  box-shadow: var(--shadow);
  font-weight: 800;
}
"""


JS = r"""const state = {
  products: [],
  filtered: [],
  quote: [],
  currentPriceFile: "",
};

const priceModes = {
  "Lista unitario": "precio_lista_unitario",
  "Neto unitario": "precio_lista_neto_unitario",
  "Neto 100 unid (prorrateado)": "precio_lista_neto_100_unid",
};

const els = {
  pickFile: document.querySelector("#pickFile"),
  priceFile: document.querySelector("#priceFile"),
  saveQuote: document.querySelector("#saveQuote"),
  clearQuote: document.querySelector("#clearQuote"),
  closeApp: document.querySelector("#closeApp"),
  statusText: document.querySelector("#statusText"),
  productCount: document.querySelector("#productCount"),
  quoteNumber: document.querySelector("#quoteNumber"),
  quoteDate: document.querySelector("#quoteDate"),
  clientName: document.querySelector("#clientName"),
  clientPhone: document.querySelector("#clientPhone"),
  clientAddress: document.querySelector("#clientAddress"),
  search: document.querySelector("#search"),
  quantity: document.querySelector("#quantity"),
  priceMode: document.querySelector("#priceMode"),
  productsBody: document.querySelector("#productsBody"),
  quoteList: document.querySelector("#quoteList"),
  emptyQuote: document.querySelector("#emptyQuote"),
  discount: document.querySelector("#discount"),
  iva: document.querySelector("#iva"),
  extra: document.querySelector("#extra"),
  subtotal: document.querySelector("#subtotal"),
  discountAmount: document.querySelector("#discountAmount"),
  ivaAmount: document.querySelector("#ivaAmount"),
  total: document.querySelector("#total"),
};

function money(value) {
  return new Intl.NumberFormat("es-AR", { style: "currency", currency: "ARS" }).format(Number(value || 0));
}

function numberValue(input) {
  const value = Number(String(input.value || "0").replace(",", "."));
  return Number.isFinite(value) && value >= 0 ? value : 0;
}

function toast(message) {
  const node = document.createElement("div");
  node.className = "toast";
  node.textContent = message;
  document.body.appendChild(node);
  window.setTimeout(() => node.remove(), 3200);
}

function fillPriceModes() {
  els.priceMode.innerHTML = Object.keys(priceModes)
    .map((name) => `<option value="${name}">${name}</option>`)
    .join("");
}

function productText(product) {
  return [product.codigo_interno, product.producto, product.descripcion_adicional, product.marca, product.indice]
    .join(" ")
    .toLowerCase()
    .normalize("NFD")
    .replace(/[\u0300-\u036f]/g, "");
}

function filterProducts() {
  const terms = els.search.value
    .trim()
    .toLowerCase()
    .normalize("NFD")
    .replace(/[\u0300-\u036f]/g, "")
    .split(/\s+/)
    .filter(Boolean);

  state.filtered = terms.length
    ? state.products.filter((product) => terms.every((term) => (product.search_text || productText(product)).includes(term)))
    : [...state.products];

  renderProducts();
}

function renderProducts() {
  const products = state.filtered.slice(0, 400);
  els.productsBody.innerHTML = products.map((product, index) => `
    <tr>
      <td>${escapeHtml(product.codigo_interno)}</td>
      <td><strong>${escapeHtml(product.producto)}</strong><br><span class="quote-meta">${escapeHtml(product.descripcion_adicional || "")}</span></td>
      <td>${escapeHtml(product.marca || "")}</td>
      <td class="money">${money(product.precio_lista_unitario)}</td>
      <td class="money">${money(product.precio_lista_neto_unitario)}</td>
      <td class="money">${money(product.precio_lista_neto_100_unid)}</td>
      <td><button class="add-row" data-index="${index}" title="Agregar">+</button></td>
    </tr>
  `).join("");

  if (!products.length) {
    els.productsBody.innerHTML = `<tr><td colspan="7">No hay productos para mostrar.</td></tr>`;
  }
}

function addProduct(product) {
  const quantity = numberValue(els.quantity);
  if (quantity <= 0) {
    toast("Ingresá una cantidad mayor que cero.");
    return;
  }

  const mode = els.priceMode.value;
  const field = priceModes[mode];
  let unitPrice = Number(product[field] || 0);
  if (field === "precio_lista_neto_100_unid") {
    unitPrice = unitPrice ? unitPrice / 100 : 0;
  }

  const detail = [product.producto, product.descripcion_adicional, product.marca].filter(Boolean).join(" | ");
  state.quote.push({
    codigo: product.codigo_interno,
    detalle: `${detail} (${mode})`,
    cantidad: quantity,
    precio_unitario: unitPrice,
  });
  renderQuote();
}

function renderQuote() {
  els.emptyQuote.classList.toggle("visible", state.quote.length === 0);
  els.quoteList.innerHTML = state.quote.map((item, index) => `
    <article class="quote-item">
      <div>
        <strong>${escapeHtml(item.detalle)}</strong>
        <div class="quote-meta">${escapeHtml(item.codigo)} · ${item.cantidad.toLocaleString("es-AR")} x ${money(item.precio_unitario)}</div>
        <div><strong>${money(item.cantidad * item.precio_unitario)}</strong></div>
      </div>
      <button class="remove" data-index="${index}" title="Quitar">×</button>
    </article>
  `).join("");
  renderTotals();
}

function renderTotals() {
  const subtotal = state.quote.reduce((sum, item) => sum + item.cantidad * item.precio_unitario, 0);
  const discountPct = numberValue(els.discount);
  const ivaPct = numberValue(els.iva);
  const extra = numberValue(els.extra);
  const discountAmount = subtotal * discountPct / 100;
  const base = subtotal - discountAmount;
  const ivaAmount = base * ivaPct / 100;
  const total = base + ivaAmount + extra;

  els.subtotal.textContent = money(subtotal);
  els.discountAmount.textContent = money(discountAmount);
  els.ivaAmount.textContent = money(ivaAmount);
  els.total.textContent = money(total);
}

function escapeHtml(value) {
  return String(value ?? "")
    .replaceAll("&", "&amp;")
    .replaceAll("<", "&lt;")
    .replaceAll(">", "&gt;")
    .replaceAll('"', "&quot;")
    .replaceAll("'", "&#039;");
}

async function uploadPriceFile(file) {
  const formData = new FormData();
  formData.append("file", file);
  els.statusText.textContent = "Cargando lista de precios...";
  const response = await fetch("/api/upload", { method: "POST", body: formData });
  const data = await response.json();
  if (!response.ok) {
    throw new Error(data.error || "No se pudo cargar la lista.");
  }
  state.products = data.products;
  state.currentPriceFile = data.current_price_file;
  filterProducts();
  els.productCount.textContent = `${state.products.length.toLocaleString("es-AR")} productos`;
  els.statusText.textContent = `Lista cargada: ${state.currentPriceFile}`;
  toast("Lista cargada correctamente.");
}

async function exportQuote() {
  if (!state.quote.length) {
    toast("Agregá al menos un producto antes de exportar.");
    return;
  }

  const payload = {
    quote_number: els.quoteNumber.value,
    date: els.quoteDate.value,
    client_name: els.clientName.value,
    client_phone: els.clientPhone.value,
    client_address: els.clientAddress.value,
    current_price_file: state.currentPriceFile,
    discount: els.discount.value,
    iva: els.iva.value,
    extra: els.extra.value,
    items: state.quote,
  };

  const response = await fetch("/api/export", {
    method: "POST",
    headers: { "Content-Type": "application/json" },
    body: JSON.stringify(payload),
  });
  if (!response.ok) {
    const data = await response.json().catch(() => ({}));
    throw new Error(data.error || "No se pudo exportar el presupuesto.");
  }

  const blob = await response.blob();
  const url = URL.createObjectURL(blob);
  const link = document.createElement("a");
  link.href = url;
  link.download = `${payload.quote_number || "presupuesto"}.xlsx`;
  document.body.appendChild(link);
  link.click();
  link.remove();
  URL.revokeObjectURL(url);
  toast("Presupuesto exportado.");
}

async function loadInitialState() {
  const response = await fetch("/api/state");
  const data = await response.json();
  state.products = data.products || [];
  state.currentPriceFile = data.current_price_file || "";
  els.quoteNumber.value = data.quote_number || "";
  els.quoteDate.value = data.date || "";
  els.productCount.textContent = `${state.products.length.toLocaleString("es-AR")} productos`;
  if (state.currentPriceFile) {
    els.statusText.textContent = `Lista cargada: ${state.currentPriceFile}`;
  }
  filterProducts();
  renderQuote();
}

els.pickFile.addEventListener("click", () => els.priceFile.click());
els.priceFile.addEventListener("change", async () => {
  const file = els.priceFile.files[0];
  if (!file) return;
  try {
    await uploadPriceFile(file);
  } catch (error) {
    els.statusText.textContent = "No se pudo cargar la lista de precios.";
    toast(error.message);
  } finally {
    els.priceFile.value = "";
  }
});

els.search.addEventListener("input", filterProducts);
els.productsBody.addEventListener("click", (event) => {
  const button = event.target.closest(".add-row");
  if (!button) return;
  addProduct(state.filtered[Number(button.dataset.index)]);
});
els.quoteList.addEventListener("click", (event) => {
  const button = event.target.closest(".remove");
  if (!button) return;
  state.quote.splice(Number(button.dataset.index), 1);
  renderQuote();
});
["discount", "iva", "extra"].forEach((id) => els[id].addEventListener("input", renderTotals));
els.clearQuote.addEventListener("click", () => {
  state.quote = [];
  els.clientName.value = "";
  els.clientPhone.value = "";
  els.clientAddress.value = "";
  els.discount.value = "0";
  els.iva.value = "21";
  els.extra.value = "0";
  renderQuote();
});
els.saveQuote.addEventListener("click", () => exportQuote().catch((error) => toast(error.message)));
els.closeApp.addEventListener("click", async () => {
  await fetch("/api/shutdown", { method: "POST" }).catch(() => {});
  document.body.innerHTML = "<div class='app-shell'><section class='panel' style='padding:24px'>Aplicación cerrada. Ya podés cerrar esta pestaña.</section></div>";
});

fillPriceModes();
loadInitialState();
"""


class RequestHandler(BaseHTTPRequestHandler):
    server_version = "PresupuestosFerreteria/1.0"

    def log_message(self, format: str, *args: object) -> None:
        return

    def do_GET(self) -> None:
        if self.path == "/" or self.path == "/index.html":
            self.send_text(HTML, "text/html; charset=utf-8")
            return
        if self.path == "/app.css":
            self.send_text(CSS, "text/css; charset=utf-8")
            return
        if self.path == "/app.js":
            self.send_text(JS, "application/javascript; charset=utf-8")
            return
        if self.path == "/api/state":
            self.send_json(STATE.snapshot())
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        if self.path == "/api/upload":
            self.handle_upload()
            return
        if self.path == "/api/export":
            self.handle_export()
            return
        if self.path == "/api/shutdown":
            self.send_json({"ok": True})
            threading.Thread(target=self.server.shutdown, daemon=True).start()
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def handle_upload(self) -> None:
        try:
            content_type = self.headers.get("Content-Type", "")
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
            )
            file_item = form["file"] if "file" in form else None
            if file_item is None or not getattr(file_item, "filename", ""):
                raise ValueError("No se recibió ningún archivo.")

            suffix = Path(file_item.filename).suffix or ".xlsx"
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as uploaded:
                uploaded.write(file_item.file.read())
                temp_path = Path(uploaded.name)

            try:
                products = load_products_from_excel(temp_path)
            finally:
                temp_path.unlink(missing_ok=True)

            STATE.set_products(products, Path(file_item.filename).name)
            self.send_json(STATE.snapshot())
        except Exception as exc:
            self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)

    def handle_export(self) -> None:
        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            quote_name = safe_file_name(str(payload.get("quote_number") or auto_quote_number()))
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as output:
                output_path = Path(output.name)
            try:
                export_quote_excel(output_path, payload)
                content = output_path.read_bytes()
            finally:
                output_path.unlink(missing_ok=True)

            file_name = f"{quote_name}.xlsx"
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", mimetypes.types_map.get(".xlsx", "application/octet-stream"))
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(file_name)}")
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
        except Exception as exc:
            self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)

    def send_text(self, content: str, content_type: str) -> None:
        encoded = content.encode("utf-8")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)

    def send_json(self, payload: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
        encoded = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)


def find_free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def load_initial_file() -> None:
    if len(sys.argv) <= 1:
        return
    candidate = Path(sys.argv[1])
    if candidate.exists():
        STATE.set_products(load_products_from_excel(candidate), candidate.name)


def main() -> None:
    load_initial_file()
    port = find_free_port()
    server = ThreadingHTTPServer(("127.0.0.1", port), RequestHandler)
    url = f"http://127.0.0.1:{port}/"
    threading.Timer(0.4, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
