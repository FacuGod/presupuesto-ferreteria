from __future__ import annotations

import math
import sys
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADER_ALIASES = {
    "codigo interno": "codigo_interno",
    "codigo interno ": "codigo_interno",
    "producto": "producto",
    "descripcion adicional": "descripcion_adicional",
    "descripcion adicional": "descripcion_adicional",
    "marca": "marca",
    "indice": "indice",
    "unidad caja granel": "unidad_caja_granel",
    "unidad caja fraccion": "unidad_caja_fraccion",
    "unidad caja fracción": "unidad_caja_fraccion",
    "precio de lista unitario": "precio_lista_unitario",
    "precio de lista neto unitario": "precio_lista_neto_unitario",
    "precio de lista neto 100 unid": "precio_lista_neto_100_unid",
}

PRICE_MODE_TO_FIELD = {
    "Lista unitario": "precio_lista_unitario",
    "Neto unitario": "precio_lista_neto_unitario",
    "Neto 100 unid (prorrateado)": "precio_lista_neto_100_unid",
}

REQUIRED_FIELDS = [
    "codigo_interno",
    "producto",
    "descripcion_adicional",
    "marca",
    "precio_lista_unitario",
    "precio_lista_neto_unitario",
    "precio_lista_neto_100_unid",
]

PALETTE = {
    "bg": "#F4F7FB",
    "surface": "#FFFFFF",
    "surface_alt": "#EEF4FF",
    "primary": "#2563EB",
    "primary_dark": "#1E40AF",
    "accent": "#F97316",
    "success": "#059669",
    "danger": "#DC2626",
    "text": "#0F172A",
    "muted": "#64748B",
    "border": "#D8E2F0",
    "row_alt": "#F8FAFC",
    "selection": "#DBEAFE",
}


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
        "\n": " ",
        "\t": " ",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return " ".join(text.split())


def to_float(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "")
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_positive_number(text: str, default: float = 0.0) -> float:
    value = to_float(text)
    return value if value >= 0 else default


def format_currency(value: float) -> str:
    formatted = f"{value:,.2f}"
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"$ {formatted}"


def auto_quote_number() -> str:
    return datetime.now().strftime("PRES-%Y%m%d-%H%M%S")


def find_header_row(ws) -> int:
    for row_index in range(1, min(ws.max_row, 40) + 1):
        row_values = [normalize_text(cell.value) for cell in ws[row_index]]
        if "codigo interno" in row_values and "producto" in row_values:
            return row_index
    raise ValueError(
        "No se encontró la fila de encabezados. La hoja debe contener columnas como 'CODIGO INTERNO' y 'PRODUCTO'."
    )


def build_header_map(header_row_values: list[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(header_row_values):
        key = HEADER_ALIASES.get(normalize_text(value))
        if key:
            mapping[key] = index
    missing = [field for field in REQUIRED_FIELDS if field not in mapping]
    if missing:
        raise ValueError(
            "Faltan columnas requeridas en el Excel: " + ", ".join(missing)
        )
    return mapping


def load_products_from_excel(file_path: str | Path) -> list[dict]:
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active
    header_row = find_header_row(sheet)
    header_values = [cell.value for cell in sheet[header_row]]
    header_map = build_header_map(header_values)

    products: list[dict] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        code = row[header_map["codigo_interno"]] if header_map["codigo_interno"] < len(row) else None
        product_name = row[header_map["producto"]] if header_map["producto"] < len(row) else None
        description = row[header_map["descripcion_adicional"]] if header_map["descripcion_adicional"] < len(row) else None
        brand = row[header_map["marca"]] if header_map["marca"] < len(row) else None

        if not any([code, product_name, description, brand]):
            continue

        item = {
            "codigo_interno": str(code).strip() if code is not None else "",
            "producto": str(product_name).strip() if product_name is not None else "",
            "descripcion_adicional": str(description).strip() if description is not None else "",
            "marca": str(brand).strip() if brand is not None else "",
            "indice": str(row[header_map.get("indice", -1)]).strip() if header_map.get("indice", -1) >= 0 and header_map.get("indice", -1) < len(row) and row[header_map.get("indice", -1)] is not None else "",
            "unidad_caja_granel": to_float(row[header_map.get("unidad_caja_granel", -1)]) if header_map.get("unidad_caja_granel", -1) >= 0 and header_map.get("unidad_caja_granel", -1) < len(row) else 0.0,
            "unidad_caja_fraccion": to_float(row[header_map.get("unidad_caja_fraccion", -1)]) if header_map.get("unidad_caja_fraccion", -1) >= 0 and header_map.get("unidad_caja_fraccion", -1) < len(row) else 0.0,
            "precio_lista_unitario": to_float(row[header_map["precio_lista_unitario"]]),
            "precio_lista_neto_unitario": to_float(row[header_map["precio_lista_neto_unitario"]]),
            "precio_lista_neto_100_unid": to_float(row[header_map["precio_lista_neto_100_unid"]]),
        }
        item["search_text"] = normalize_text(
            " ".join(
                [
                    item["codigo_interno"],
                    item["producto"],
                    item["descripcion_adicional"],
                    item["marca"],
                    item["indice"],
                ]
            )
        )
        products.append(item)

    if not products:
        raise ValueError("El archivo se abrió, pero no se encontraron productos debajo del encabezado.")
    return products


class FerreteriaBudgetApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Presupuestos - Ferretería")
        self.geometry("1460x860")
        self.minsize(1200, 720)

        self.products: list[dict] = []
        self.filtered_products: list[dict] = []
        self.quote_items: list[dict] = []
        self.current_price_file: str = ""

        self.search_var = tk.StringVar()
        self.quantity_var = tk.StringVar(value="1")
        self.price_mode_var = tk.StringVar(value="Lista unitario")
        self.quote_number_var = tk.StringVar(value=auto_quote_number())
        self.date_var = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y"))
        self.client_name_var = tk.StringVar()
        self.client_phone_var = tk.StringVar()
        self.client_address_var = tk.StringVar()
        self.discount_var = tk.StringVar(value="0")
        self.iva_var = tk.StringVar(value="21")
        self.extra_var = tk.StringVar(value="0")
        self.status_var = tk.StringVar(value="Cargá una lista de precios para empezar.")

        self.subtotal_value_var = tk.StringVar(value=format_currency(0))
        self.discount_value_var = tk.StringVar(value=format_currency(0))
        self.iva_value_var = tk.StringVar(value=format_currency(0))
        self.total_value_var = tk.StringVar(value=format_currency(0))

        self._configure_theme()
        self._build_ui()
        self._bind_events()

        if len(sys.argv) > 1:
            candidate = Path(sys.argv[1])
            if candidate.exists():
                self.load_price_list(candidate)

    def _configure_theme(self) -> None:
        self.configure(bg=PALETTE["bg"])
        self.option_add("*Font", ("Segoe UI", 10))

        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(".", font=("Segoe UI", 10), background=PALETTE["bg"], foreground=PALETTE["text"])
        style.configure("App.TFrame", background=PALETTE["bg"])
        style.configure("Surface.TFrame", background=PALETTE["surface"])
        style.configure("Panel.TLabelframe", background=PALETTE["surface"], bordercolor=PALETTE["border"], relief="solid")
        style.configure(
            "Panel.TLabelframe.Label",
            background=PALETTE["surface"],
            foreground=PALETTE["primary_dark"],
            font=("Segoe UI", 11, "bold"),
        )
        style.configure("TLabel", background=PALETTE["surface"], foreground=PALETTE["text"])
        style.configure("Muted.TLabel", background=PALETTE["surface"], foreground=PALETTE["muted"])
        style.configure("Hero.TLabel", background=PALETTE["primary_dark"], foreground="#FFFFFF", font=("Segoe UI", 20, "bold"))
        style.configure("HeroSub.TLabel", background=PALETTE["primary_dark"], foreground="#BFDBFE", font=("Segoe UI", 10))
        style.configure("Status.TLabel", background="#DBEAFE", foreground=PALETTE["primary_dark"], padding=(12, 7))
        style.configure("TotalLabel.TLabel", background=PALETTE["surface"], foreground=PALETTE["muted"], font=("Segoe UI", 10, "bold"))
        style.configure("TotalValue.TLabel", background=PALETTE["primary"], foreground="#FFFFFF", font=("Segoe UI", 18, "bold"), padding=(14, 8))

        style.configure("TEntry", fieldbackground="#FFFFFF", bordercolor=PALETTE["border"], lightcolor=PALETTE["border"], padding=7)
        style.map("TEntry", bordercolor=[("focus", PALETTE["primary"])])
        style.configure("TCombobox", fieldbackground="#FFFFFF", bordercolor=PALETTE["border"], padding=6)

        style.configure(
            "Primary.TButton",
            background=PALETTE["primary"],
            foreground="#FFFFFF",
            bordercolor=PALETTE["primary"],
            focusthickness=0,
            padding=(14, 9),
            font=("Segoe UI", 10, "bold"),
        )
        style.map("Primary.TButton", background=[("active", PALETTE["primary_dark"]), ("pressed", PALETTE["primary_dark"])])
        style.configure(
            "Accent.TButton",
            background=PALETTE["accent"],
            foreground="#FFFFFF",
            bordercolor=PALETTE["accent"],
            focusthickness=0,
            padding=(14, 9),
            font=("Segoe UI", 10, "bold"),
        )
        style.map("Accent.TButton", background=[("active", "#EA580C"), ("pressed", "#C2410C")])
        style.configure(
            "Danger.TButton",
            background="#FEE2E2",
            foreground=PALETTE["danger"],
            bordercolor="#FCA5A5",
            focusthickness=0,
            padding=(12, 8),
            font=("Segoe UI", 10, "bold"),
        )
        style.map("Danger.TButton", background=[("active", "#FECACA"), ("pressed", "#FCA5A5")])
        style.configure("Secondary.TButton", background="#E0F2FE", foreground="#075985", bordercolor="#BAE6FD", padding=(12, 8))
        style.map("Secondary.TButton", background=[("active", "#BAE6FD"), ("pressed", "#7DD3FC")])

        style.configure(
            "Treeview",
            background=PALETTE["surface"],
            fieldbackground=PALETTE["surface"],
            foreground=PALETTE["text"],
            rowheight=30,
            bordercolor=PALETTE["border"],
            borderwidth=0,
        )
        style.configure(
            "Treeview.Heading",
            background=PALETTE["primary_dark"],
            foreground="#FFFFFF",
            relief="flat",
            font=("Segoe UI", 10, "bold"),
            padding=(8, 8),
        )
        style.map("Treeview", background=[("selected", PALETTE["selection"])], foreground=[("selected", PALETTE["text"])])
        style.map("Treeview.Heading", background=[("active", PALETTE["primary"])])
        style.configure("TPanedwindow", background=PALETTE["bg"])

    def _field(self, parent: tk.Widget, label: str, variable: tk.StringVar, row: int, column: int, width: int = 20, columnspan: int = 1) -> None:
        ttk.Label(parent, text=label, style="Muted.TLabel").grid(row=row, column=column, sticky="w", padx=8, pady=(5, 2))
        ttk.Entry(parent, textvariable=variable, width=width).grid(
            row=row + 1,
            column=column,
            columnspan=columnspan,
            sticky="ew",
            padx=8,
            pady=(0, 10),
        )

    def _metric(self, parent: tk.Widget, label: str, variable: tk.StringVar, row: int, column: int) -> None:
        ttk.Label(parent, text=label, style="TotalLabel.TLabel").grid(row=row, column=column, sticky="w", padx=8, pady=(7, 2))
        ttk.Label(parent, textvariable=variable).grid(row=row + 1, column=column, sticky="ew", padx=8, pady=(0, 8))

    def _build_ui(self) -> None:
        hero = tk.Frame(self, bg=PALETTE["primary_dark"], padx=22, pady=16)
        hero.pack(fill="x")

        title_area = tk.Frame(hero, bg=PALETTE["primary_dark"])
        title_area.pack(side="left", fill="x", expand=True)
        ttk.Label(title_area, text="Presupuestos Ferretería", style="Hero.TLabel").pack(anchor="w")
        ttk.Label(title_area, text="Carga listas, arma presupuestos y exporta Excel desde una sola pantalla.", style="HeroSub.TLabel").pack(anchor="w", pady=(3, 0))

        action_area = tk.Frame(hero, bg=PALETTE["primary_dark"])
        action_area.pack(side="right")
        ttk.Button(action_area, text="Cargar lista", style="Accent.TButton", command=self.select_price_list).pack(side="left", padx=(0, 8))
        ttk.Button(action_area, text="Guardar Excel", style="Primary.TButton", command=self.save_quote_to_excel).pack(side="left", padx=(0, 8))
        ttk.Button(action_area, text="Limpiar", style="Secondary.TButton", command=self.clear_quote).pack(side="left")

        status_frame = ttk.Frame(self, style="App.TFrame", padding=(16, 12, 16, 0))
        status_frame.pack(fill="x")
        ttk.Label(status_frame, textvariable=self.status_var, style="Status.TLabel").pack(fill="x")

        info_frame = ttk.LabelFrame(self, text="Datos del presupuesto", style="Panel.TLabelframe", padding=14)
        info_frame.pack(fill="x", padx=16, pady=(12, 10))

        self._field(info_frame, "Número", self.quote_number_var, 0, 0, width=26)
        self._field(info_frame, "Fecha", self.date_var, 0, 1, width=16)
        self._field(info_frame, "Cliente", self.client_name_var, 0, 2, width=34)
        self._field(info_frame, "Teléfono", self.client_phone_var, 0, 3, width=20)
        self._field(info_frame, "Dirección / Observación", self.client_address_var, 2, 0, width=50, columnspan=4)

        for column in range(4):
            info_frame.columnconfigure(column, weight=1)

        body = ttk.PanedWindow(self, orient="horizontal")
        body.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        left_panel = ttk.Frame(body, style="App.TFrame", padding=(0, 0, 8, 0))
        right_panel = ttk.Frame(body, style="App.TFrame", padding=(8, 0, 0, 0))
        body.add(left_panel, weight=5)
        body.add(right_panel, weight=4)

        search_box = ttk.LabelFrame(left_panel, text="Productos", style="Panel.TLabelframe", padding=14)
        search_box.pack(fill="both", expand=True)

        controls = ttk.Frame(search_box, style="Surface.TFrame")
        controls.pack(fill="x", pady=(0, 12))
        ttk.Label(controls, text="Buscar", style="Muted.TLabel").pack(side="left")
        ttk.Entry(controls, textvariable=self.search_var, width=46).pack(side="left", padx=8, fill="x", expand=True)
        ttk.Label(controls, text="Cantidad", style="Muted.TLabel").pack(side="left", padx=(10, 0))
        ttk.Entry(controls, textvariable=self.quantity_var, width=8).pack(side="left", padx=8)
        ttk.Label(controls, text="Precio", style="Muted.TLabel").pack(side="left")
        ttk.Combobox(
            controls,
            textvariable=self.price_mode_var,
            values=list(PRICE_MODE_TO_FIELD.keys()),
            width=28,
            state="readonly",
        ).pack(side="left", padx=6)
        ttk.Button(controls, text="Agregar", style="Primary.TButton", command=self.add_selected_product).pack(side="left", padx=(8, 0))

        product_columns = (
            "codigo",
            "producto",
            "descripcion",
            "marca",
            "precio_lista",
            "precio_neto",
            "precio_100",
        )
        product_table = ttk.Frame(search_box, style="Surface.TFrame")
        product_table.pack(fill="both", expand=True)

        self.product_tree = ttk.Treeview(product_table, columns=product_columns, show="headings", height=22)
        headings = {
            "codigo": "Código",
            "producto": "Producto",
            "descripcion": "Descripción",
            "marca": "Marca",
            "precio_lista": "Lista",
            "precio_neto": "Neto unit.",
            "precio_100": "Neto 100",
        }
        widths = {
            "codigo": 100,
            "producto": 320,
            "descripcion": 150,
            "marca": 140,
            "precio_lista": 95,
            "precio_neto": 95,
            "precio_100": 95,
        }
        for col in product_columns:
            self.product_tree.heading(col, text=headings[col])
            anchor = "e" if "precio" in col else "w"
            self.product_tree.column(col, width=widths[col], anchor=anchor)

        product_scroll_y = ttk.Scrollbar(product_table, orient="vertical", command=self.product_tree.yview)
        product_scroll_x = ttk.Scrollbar(product_table, orient="horizontal", command=self.product_tree.xview)
        self.product_tree.configure(yscrollcommand=product_scroll_y.set, xscrollcommand=product_scroll_x.set)
        self.product_tree.tag_configure("odd", background=PALETTE["surface"])
        self.product_tree.tag_configure("even", background=PALETTE["row_alt"])
        self.product_tree.grid(row=0, column=0, sticky="nsew")
        product_scroll_y.grid(row=0, column=1, sticky="ns")
        product_scroll_x.grid(row=1, column=0, sticky="ew")
        product_table.columnconfigure(0, weight=1)
        product_table.rowconfigure(0, weight=1)

        quote_box = ttk.LabelFrame(right_panel, text="Presupuesto actual", style="Panel.TLabelframe", padding=14)
        quote_box.pack(fill="both", expand=True)

        quote_columns = ("item", "codigo", "detalle", "cantidad", "precio_unitario", "subtotal")
        quote_table = ttk.Frame(quote_box, style="Surface.TFrame")
        quote_table.pack(fill="both", expand=True)

        self.quote_tree = ttk.Treeview(quote_table, columns=quote_columns, show="headings", height=20)
        quote_headings = {
            "item": "#",
            "codigo": "Código",
            "detalle": "Detalle",
            "cantidad": "Cant.",
            "precio_unitario": "Precio unit.",
            "subtotal": "Subtotal",
        }
        quote_widths = {
            "item": 45,
            "codigo": 100,
            "detalle": 360,
            "cantidad": 70,
            "precio_unitario": 110,
            "subtotal": 120,
        }
        for col in quote_columns:
            self.quote_tree.heading(col, text=quote_headings[col])
            anchor = "e" if col in {"cantidad", "precio_unitario", "subtotal"} else "w"
            self.quote_tree.column(col, width=quote_widths[col], anchor=anchor)

        quote_scroll_y = ttk.Scrollbar(quote_table, orient="vertical", command=self.quote_tree.yview)
        self.quote_tree.configure(yscrollcommand=quote_scroll_y.set)
        self.quote_tree.tag_configure("odd", background=PALETTE["surface"])
        self.quote_tree.tag_configure("even", background=PALETTE["row_alt"])
        self.quote_tree.grid(row=0, column=0, sticky="nsew")
        quote_scroll_y.grid(row=0, column=1, sticky="ns")
        quote_table.columnconfigure(0, weight=1)
        quote_table.rowconfigure(0, weight=1)

        actions_frame = ttk.Frame(quote_box, style="Surface.TFrame")
        actions_frame.pack(fill="x", pady=(10, 0))
        ttk.Button(actions_frame, text="Eliminar seleccionado", style="Danger.TButton", command=self.remove_selected_quote_item).pack(side="left")

        totals_frame = ttk.LabelFrame(right_panel, text="Cálculo final", style="Panel.TLabelframe", padding=14)
        totals_frame.pack(fill="x", pady=(12, 0))

        ttk.Label(totals_frame, text="Descuento %", style="Muted.TLabel").grid(row=0, column=0, sticky="w", padx=8, pady=(4, 2))
        ttk.Entry(totals_frame, textvariable=self.discount_var, width=10).grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 10))
        ttk.Label(totals_frame, text="IVA %", style="Muted.TLabel").grid(row=0, column=1, sticky="w", padx=8, pady=(4, 2))
        ttk.Entry(totals_frame, textvariable=self.iva_var, width=10).grid(row=1, column=1, sticky="ew", padx=8, pady=(0, 10))
        ttk.Label(totals_frame, text="Recargo / Flete", style="Muted.TLabel").grid(row=0, column=2, sticky="w", padx=8, pady=(4, 2))
        ttk.Entry(totals_frame, textvariable=self.extra_var, width=12).grid(row=1, column=2, sticky="ew", padx=8, pady=(0, 10))

        self._metric(totals_frame, "Subtotal productos", self.subtotal_value_var, 2, 0)
        self._metric(totals_frame, "Descuento", self.discount_value_var, 2, 1)
        self._metric(totals_frame, "IVA", self.iva_value_var, 2, 2)
        ttk.Label(totals_frame, text="TOTAL", style="TotalLabel.TLabel").grid(row=4, column=0, sticky="w", padx=8, pady=(8, 2))
        ttk.Label(totals_frame, textvariable=self.total_value_var, style="TotalValue.TLabel").grid(row=5, column=0, columnspan=3, sticky="ew", padx=8, pady=(0, 4))

        for column in range(3):
            totals_frame.columnconfigure(column, weight=1)

    def _bind_events(self) -> None:
        self.search_var.trace_add("write", lambda *_: self.refresh_product_table())
        self.discount_var.trace_add("write", lambda *_: self.refresh_totals())
        self.iva_var.trace_add("write", lambda *_: self.refresh_totals())
        self.extra_var.trace_add("write", lambda *_: self.refresh_totals())
        self.product_tree.bind("<Double-1>", lambda _event: self.add_selected_product())

    def select_price_list(self) -> None:
        file_path = filedialog.askopenfilename(
            title="Seleccionar lista de precios",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xltx *.xltm")],
        )
        if file_path:
            self.load_price_list(file_path)

    def load_price_list(self, file_path: str | Path) -> None:
        try:
            self.products = load_products_from_excel(file_path)
            self.filtered_products = list(self.products)
            self.current_price_file = str(file_path)
            self.refresh_product_table()
            self.status_var.set(
                f"Lista cargada: {Path(file_path).name} | Productos detectados: {len(self.products)}"
            )
        except Exception as exc:
            messagebox.showerror("Error al abrir archivo", str(exc))
            self.status_var.set("No se pudo abrir la lista de precios.")

    def refresh_product_table(self) -> None:
        query = normalize_text(self.search_var.get())
        self.product_tree.delete(*self.product_tree.get_children())

        if not self.products:
            return

        if query:
            terms = query.split()
            self.filtered_products = [
                product
                for product in self.products
                if all(term in product["search_text"] for term in terms)
            ]
        else:
            self.filtered_products = list(self.products)

        for index, product in enumerate(self.filtered_products):
            self.product_tree.insert(
                "",
                "end",
                iid=str(index),
                tags=("even" if index % 2 == 0 else "odd",),
                values=(
                    product["codigo_interno"],
                    product["producto"],
                    product["descripcion_adicional"],
                    product["marca"],
                    format_currency(product["precio_lista_unitario"]),
                    format_currency(product["precio_lista_neto_unitario"]),
                    format_currency(product["precio_lista_neto_100_unid"]),
                ),
            )

    def get_selected_product(self) -> dict | None:
        selected = self.product_tree.selection()
        if not selected:
            return None
        selected_index = int(selected[0])
        if selected_index >= len(self.filtered_products):
            return None
        return self.filtered_products[selected_index]

    def add_selected_product(self) -> None:
        if not self.products:
            messagebox.showwarning("Sin lista", "Primero tenés que cargar una lista de precios.")
            return

        product = self.get_selected_product()
        if product is None:
            messagebox.showinfo("Seleccionar producto", "Elegí un producto para agregar al presupuesto.")
            return

        quantity = parse_positive_number(self.quantity_var.get(), default=-1)
        if quantity <= 0:
            messagebox.showwarning("Cantidad inválida", "Ingresá una cantidad mayor que cero.")
            return

        price_mode = self.price_mode_var.get()
        price_field = PRICE_MODE_TO_FIELD[price_mode]
        unit_price = product.get(price_field, 0.0)
        if price_field == "precio_lista_neto_100_unid":
            unit_price = unit_price / 100 if unit_price else 0.0

        detail_parts = [product["producto"], product["descripcion_adicional"], product["marca"]]
        detail = " | ".join(part for part in detail_parts if part)

        item = {
            "codigo": product["codigo_interno"],
            "detalle": detail,
            "cantidad": quantity,
            "precio_unitario": unit_price,
            "subtotal": quantity * unit_price,
            "modo_precio": price_mode,
        }
        self.quote_items.append(item)
        self.refresh_quote_table()
        self.refresh_totals()

    def refresh_quote_table(self) -> None:
        self.quote_tree.delete(*self.quote_tree.get_children())
        for index, item in enumerate(self.quote_items, start=1):
            self.quote_tree.insert(
                "",
                "end",
                iid=str(index - 1),
                tags=("even" if index % 2 == 0 else "odd",),
                values=(
                    index,
                    item["codigo"],
                    f"{item['detalle']} ({item['modo_precio']})",
                    f"{item['cantidad']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                    format_currency(item["precio_unitario"]),
                    format_currency(item["subtotal"]),
                ),
            )

    def remove_selected_quote_item(self) -> None:
        selected = self.quote_tree.selection()
        if not selected:
            return
        index = int(selected[0])
        if 0 <= index < len(self.quote_items):
            self.quote_items.pop(index)
            self.refresh_quote_table()
            self.refresh_totals()

    def refresh_totals(self) -> None:
        subtotal = sum(item["subtotal"] for item in self.quote_items)
        discount_pct = parse_positive_number(self.discount_var.get())
        iva_pct = parse_positive_number(self.iva_var.get())
        extra_cost = parse_positive_number(self.extra_var.get())

        discount_amount = subtotal * discount_pct / 100
        base = subtotal - discount_amount
        iva_amount = base * iva_pct / 100
        total = base + iva_amount + extra_cost

        self.subtotal_value_var.set(format_currency(subtotal))
        self.discount_value_var.set(format_currency(discount_amount))
        self.iva_value_var.set(format_currency(iva_amount))
        self.total_value_var.set(format_currency(total))

    def clear_quote(self) -> None:
        self.quote_items.clear()
        self.quote_number_var.set(auto_quote_number())
        self.client_name_var.set("")
        self.client_phone_var.set("")
        self.client_address_var.set("")
        self.discount_var.set("0")
        self.extra_var.set("0")
        self.refresh_quote_table()
        self.refresh_totals()

    def save_quote_to_excel(self) -> None:
        if not self.quote_items:
            messagebox.showinfo("Sin ítems", "Agregá al menos un producto antes de guardar el presupuesto.")
            return

        default_name = f"{self.quote_number_var.get().strip() or auto_quote_number()}.xlsx"
        file_path = filedialog.asksaveasfilename(
            title="Guardar presupuesto",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not file_path:
            return

        try:
            self._export_quote_excel(file_path)
            messagebox.showinfo("Listo", f"Presupuesto guardado en:\n{file_path}")
        except Exception as exc:
            messagebox.showerror("Error al guardar", str(exc))

    def _export_quote_excel(self, file_path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Presupuesto"

        bold = Font(bold=True)
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin_side = Side(style="thin", color="B7C9D6")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        ws["A1"] = "PRESUPUESTO"
        ws["A1"].font = title_font
        ws.merge_cells("A1:F1")
        ws["A1"].alignment = Alignment(horizontal="center")

        metadata = [
            ("Número", self.quote_number_var.get().strip()),
            ("Fecha", self.date_var.get().strip()),
            ("Cliente", self.client_name_var.get().strip()),
            ("Teléfono", self.client_phone_var.get().strip()),
            ("Dirección / Observación", self.client_address_var.get().strip()),
            ("Lista de precios usada", self.current_price_file),
        ]
        row_cursor = 3
        for label, value in metadata:
            ws.cell(row=row_cursor, column=1, value=label).font = bold
            ws.cell(row=row_cursor, column=2, value=value)
            row_cursor += 1

        row_cursor += 1
        headers = ["#", "Código", "Detalle", "Cantidad", "Precio unitario", "Subtotal"]
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_cursor, column=col_index, value=header)
            cell.font = bold
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        data_start = row_cursor + 1
        for item_index, item in enumerate(self.quote_items, start=1):
            row_idx = data_start + item_index - 1
            values = [
                item_index,
                item["codigo"],
                f"{item['detalle']} ({item['modo_precio']})",
                item["cantidad"],
                item["precio_unitario"],
                item["subtotal"],
            ]
            for col_index, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_index, value=value)
                cell.border = border
                if col_index in {4, 5, 6}:
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        for row_idx in range(data_start, data_start + len(self.quote_items)):
            ws.cell(row=row_idx, column=4).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=5).number_format = '$ #,##0.00'
            ws.cell(row=row_idx, column=6).number_format = '$ #,##0.00'

        subtotal = sum(item["subtotal"] for item in self.quote_items)
        discount_pct = parse_positive_number(self.discount_var.get())
        iva_pct = parse_positive_number(self.iva_var.get())
        extra_cost = parse_positive_number(self.extra_var.get())
        discount_amount = subtotal * discount_pct / 100
        base = subtotal - discount_amount
        iva_amount = base * iva_pct / 100
        total = base + iva_amount + extra_cost

        total_start = data_start + len(self.quote_items) + 2
        totals = [
            ("Subtotal", subtotal),
            (f"Descuento ({discount_pct:.2f}%)", discount_amount),
            (f"IVA ({iva_pct:.2f}%)", iva_amount),
            ("Recargo / Flete", extra_cost),
            ("TOTAL", total),
        ]
        for offset, (label, value) in enumerate(totals):
            row_idx = total_start + offset
            ws.cell(row=row_idx, column=5, value=label).font = bold
            amount_cell = ws.cell(row=row_idx, column=6, value=value)
            amount_cell.font = bold if label == "TOTAL" else Font(bold=False)
            amount_cell.number_format = '$ #,##0.00'
            ws.cell(row=row_idx, column=5).border = border
            ws.cell(row=row_idx, column=6).border = border

        widths = {
            "A": 8,
            "B": 16,
            "C": 60,
            "D": 12,
            "E": 18,
            "F": 18,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows():
            for cell in row:
                if cell.column == 3:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(file_path)


def main() -> None:
    app = FerreteriaBudgetApp()
    app.mainloop()


if __name__ == "__main__":
    main()
